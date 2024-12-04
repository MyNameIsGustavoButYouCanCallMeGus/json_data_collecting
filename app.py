from flask import Flask, request, render_template, send_file
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import os

app = Flask(__name__)

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/sum', methods=['GET', 'POST'])
def index():
    dates = [
        {"date": "2022", "file": "json_data/all_data_2022.json"},
        {"date": "2023", "file": "json_data/all_data_2023.json"},
        {"date": "На январь", "file": "json_data/all_data_january.json"},
        {"date": "На февраль", "file": "json_data/all_data_february.json"},
        {"date": "На март", "file": "json_data/all_data_march.json"},
        {"date": "На апрель", "file": "json_data/all_data_april.json"},
        {"date": "На май", "file": "json_data/all_data_may.json"},
        {"date": "На июнь", "file": "json_data/all_data_june.json"},
    ]

    form_data = {
        'grbId': '',
        'katFgr': '',
        'klsFpgr': '',
        'pklAbp': '',
        'katoName':'',
        'kato': ''
    }

    total_obz = 0
    katFgrName_list = []

    if request.method == 'POST':
        form_data['grbId'] = request.form.get('grbId', '')
        form_data['katFgr'] = request.form.get('katFgr', '')
        form_data['klsFpgr'] = request.form.get('klsFpgr', '')
        form_data['pklAbp'] = request.form.get('pklAbp', '')
        form_data['katoName'] = request.form.get('katoName', '')
        form_data['kato'] = request.form.get('kato', '')

        file_name = request.form.get('file')
        if file_name and os.path.exists(file_name):
            with open(file_name, 'r', encoding='utf-8') as file:
                all_data = json.load(file)

            for entry in all_data:
                if (
                        (not form_data['grbId'] or entry.get('grbId') == int(form_data['grbId'])) and
                        (not form_data['katFgr'] or entry.get('katFgr') == form_data['katFgr']) and
                        (not form_data['klsFpgr'] or entry.get('klsFpgr') == form_data['klsFpgr']) and
                        (not form_data['pklAbp'] or entry.get('pklAbp') == form_data['pklAbp']) and
                        (not form_data['katoName'] or entry.get('katoName') == form_data['katoName']) and
                        (not form_data['kato'] or entry.get('kato') == form_data['kato'])
                ):
                    katFgrName_list.append(entry.get('katFgrName'))
                    obz_value_str = entry.get('sumrg', '0').replace(',', '')
                    obz_value = round(float(obz_value_str) / 1_000_000) if obz_value_str else 0.0
                    total_obz += obz_value
        else:
            return "File not found or not specified", 404

    return render_template('index.html', form_data=form_data, katFgrName_list=katFgrName_list, total_obz=total_obz, dates=dates)


@app.route('/search', methods=['GET', 'POST'])
def search():
    results = []

    dates = [
        {"date": "2022", "file": "json_data/all_data_2022.json"},
        {"date": "2023", "file": "json_data/all_data_2023.json"},
        {"date": "На январь", "file": "json_data/all_data_january.json"},
        {"date": "На февраль", "file": "json_data/all_data_february.json"},
        {"date": "На март", "file": "json_data/all_data_march.json"},
        {"date": "На апрель", "file": "json_data/all_data_april.json"},
        {"date": "На май", "file": "json_data/all_data_may.json"},
        {"date": "На июнь", "file": "json_data/all_data_june.json"},
    ]

    form_data = {
        'grbId': '',
        'katFgrName': '',
        'klsFpgrName': '',
        'pklAbpName': '',
        'katoName': '',
        'kato': ''
    }

    if request.method == 'POST':
        form_data['grbId'] = request.form.get('grbId', '')
        form_data['katFgrName'] = request.form.get('katFgrName', '')
        form_data['klsFpgrName'] = request.form.get('klsFpgrName', '')
        form_data['pklAbpName'] = request.form.get('pklAbpName', '')
        form_data['katoName'] = request.form.get('katoName', '')
        form_data['kato'] = request.form.get('kato', '')

        file_name = request.form.get('file')
        if file_name and os.path.exists(file_name):
            with open(file_name, 'r', encoding='utf-8') as file:
                all_data = json.load(file)

            for entry in all_data:
                if (
                        (not form_data['grbId'] or entry.get('grbId') == int(form_data['grbId'])) and
                        (not form_data['katFgrName'] or entry.get('katFgrName') == form_data['katFgrName']) and
                        (not form_data['klsFpgrName'] or entry.get('klsFpgrName') == form_data['klsFpgrName']) and
                        (not form_data['pklAbpName'] or entry.get('pklAbpName') == form_data['pklAbpName']) and
                        (not form_data['katoName'] or entry.get('katoName') == form_data['katoName']) and
                        (not form_data['kato'] or entry.get('kato') == form_data['kato'])
                ):
                    results.append(entry)
        else:
            return "File not found or not specified", 404

    return render_template('search.html', form_data=form_data, results=results, dates=dates)


def create_table_in_excel(data, region_name, output_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    font_size = 9
    font_size2 = 7

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12.5

    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='center')
            cell.font = Font(name='Times New Roman', size=font_size)

    ws.merge_cells('A1:H1')
    ws['A1'].value = 'Отчет об исполнении бюджета'
    ws['A1'].font = Font(name='Times New Roman', bold=True)
    date_1 = request.form.get('date')
    dates = [
        {"date": "2022", "file": "json_data/all_data_2022.json"},
        {"date": "2023", "file": "json_data/all_data_2023.json"},
        {"date": "До января", "file": "json_data/all_data_january.json"},
        {"date": "До февраля", "file": "json_data/all_data_february.json"},
        {"date": "До марта", "file": "json_data/all_data_march.json"},
        {"date": "До апреля", "file": "json_data/all_data_april.json"},
        {"date": "До мая", "file": "json_data/all_data_may.json"},
        {"date": "До июня", "file": "json_data/all_data_june.json"},
    ]

    date_name = ""
    for i in dates:
        if date_1 == i["file"]:
            date_name = i["date"]
            break

    ws.merge_cells('A2:H2')
    ws['A2'].value = date_name + ' 2024 года'
    ws['A2'].font = Font(name='Times New Roman', bold=True)

    ws.merge_cells('A4:D4')
    ws['A4'].value = 'Индекс'
    ws['A4'].font = Font(name='Times New Roman', size=font_size)
    ws.merge_cells('E4:H4')
    ws['E4'].value = 'форма 7-ОИБ'
    ws['E4'].font = Font(name='Times New Roman', size=font_size)

    ws.merge_cells('A5:D5')
    ws['A5'].value = 'Круг лиц, представляющих:'
    ws['A5'].font = Font(name='Times New Roman', size=font_size)
    ws.merge_cells('E5:H5')
    ws['E5'].value = 'местные уполномоченные органы по исполнению бюджета'
    ws['E5'].font = Font(name='Times New Roman', size=font_size)

    ws.merge_cells('A6:D6')
    ws['A6'].value = 'Куда представляется:'
    ws['A6'].font = Font(name='Times New Roman', size=font_size)
    ws.merge_cells('E6:H6')
    ws['E6'].value = 'в уполномоченный орган по исполнению вышестоящего бюджета'
    ws['E6'].font = Font(name='Times New Roman', size=font_size)

    ws.merge_cells('A7:D7')
    ws['A7'].value = 'Период:'
    ws['A7'].font = Font(name='Times New Roman', size=font_size)
    ws.merge_cells('E7:H7')
    ws['E7'].value = date_name.lower()
    ws['E7'].font = Font(name='Times New Roman', size=font_size)

    ws.merge_cells('A8:D8')
    ws['A8'].value = 'Срок представления:'
    ws['A8'].font = Font(name='Times New Roman', size=font_size)
    ws.row_dimensions[8].height = 40
    ws.row_dimensions[11].height = 5
    ws.merge_cells('E8:H8')
    ws['E8'].value = 'для аппаратов акимов городов районного значения, сел, поселков, сельских округов устанавливаются уполномоченными органами по исполнению бюджета района (города областного значения)'
    ws['E8'].font = Font(name='Times New Roman', size=font_size)
    ws['E8'].alignment = Alignment(wrap_text=True)

    ws.merge_cells('A9:D9')
    ws['A9'].value = 'Регион:'
    ws['A9'].font = Font(name='Times New Roman', size=font_size)
    ws.merge_cells('E9:H9')
    ws['E9'].value = region_name
    ws['E9'].font = Font(name='Times New Roman', size=font_size)
    ws['E9'].alignment = Alignment(wrap_text=True)

    ws.merge_cells('A10:G10')
    ws['H10'].value = 'тыс. тенге'
    ws['H10'].font = Font(name='Times New Roman', bold=True, size=font_size)
    ws.merge_cells('A11:H11')

    headers = ["Исполнение поступлениий бюджета и/или оплаченных обязательств по бюджетным программам (подпрограммам)"]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=8)
        cell.value = header
        cell.font = Font(name='Times New Roman', size=font_size)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers2 = ["Наименование"]

    for col_num2, header2 in enumerate(headers2, start=1):
        cell = ws.cell(row=12, column=7)
        cell.value = header2
        cell.font = Font(name='Times New Roman', size=font_size)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A14:F14')
    ws['A14'].value = '1'
    ws['G14'].value = '2'
    ws['G14'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H14'].value = '3'
    ws['H14'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A12:F13')
    ws['A12'].value = 'Коды бюджетной классификации'

    headers3 = ["I. Доходы"]

    for col_num3, header3 in enumerate(headers3, start=1):
        cell = ws.cell(row=15, column=7)
        cell.value = header3.upper()
        cell.font = Font(name='Times New Roman', bold=True, size=font_size)

    headers3_1 = ["1"]

    for col_num3_1, header3_1 in enumerate(headers3_1, start=1):
        cell = ws.cell(row=15, column=1)
        cell.value = header3_1.upper()
        cell.font = Font(name='Times New Roman', size=font_size)

    aggregated_data_income = {}
    total_sum_income = 0.0
    for item in data:
        if item.get('grbId') == 1:
            katFgr = item.get('katFgr', '')
            katFgrName = item.get('katFgrName', '')
            klsFpgrName = item.get('klsFpgrName', '')
            sumrg_str = item.get('sumrg', '0').replace(',', '.')
            try:
                sumrg = round(float(sumrg_str) / 1_000_000, 1) if sumrg_str else 0.0
                total_sum_income += sumrg
            except ValueError:
                sumrg = 0.0

            if katFgr not in aggregated_data_income:
                aggregated_data_income[katFgr] = {
                    'katFgrName': katFgrName,
                    'totalSum': 0.0,
                    'klsFpgrNames': {}
                }

            if klsFpgrName in aggregated_data_income[katFgr]['klsFpgrNames']:
                aggregated_data_income[katFgr]['klsFpgrNames'][klsFpgrName] += sumrg
            else:
                aggregated_data_income[katFgr]['klsFpgrNames'][klsFpgrName] = sumrg

            aggregated_data_income[katFgr]['totalSum'] += sumrg

    sorted_data_income = dict(sorted(aggregated_data_income.items()))

    row_num = 16
    counter = 1
    counter_in = 1
    for katFgr, katFgr_data in sorted_data_income.items():
        cell0 = ws.cell(row=row_num, column=3)
        cell1 = ws.cell(row=row_num, column=7)
        cell2 = ws.cell(row=row_num, column=8)
        cell0.value = counter
        cell1.value = katFgr_data['katFgrName'].upper()
        cell2.value = katFgr_data['totalSum']
        cell0.font = Font(name='Times New Roman', size=font_size)
        cell1.font = Font(name='Times New Roman', bold=True, size=font_size)
        cell2.font = Font(name='Times New Roman', bold=True, size=font_size)
        row_num += 1
        counter += 1

        for klsFpgrName, sumrg in katFgr_data['klsFpgrNames'].items():
            cell0 = ws.cell(row=row_num, column=5)
            cell1 = ws.cell(row=row_num, column=7)
            cell2 = ws.cell(row=row_num, column=8)
            cell0.value = counter_in
            cell1.value = klsFpgrName
            cell2.value = sumrg
            cell0.font = Font(name='Times New Roman', size=font_size2)
            cell1.font = Font(name='Times New Roman', size=font_size2)
            cell2.font = Font(name='Times New Roman', size=font_size)
            cell1.alignment = Alignment(wrap_text=True)
            row_num += 1
            counter_in += 1

    total_sum_cell_income = ws.cell(row=15, column=8)
    total_sum_cell_income.value = total_sum_income
    total_sum_cell_income.font = Font(name='Times New Roman', bold=True, size=font_size)

    # Add a header for expenses
    row_num += 0

    headers4 = ["II. Затраты"]

    for col_num4, header4 in enumerate(headers4, start=1):
        cell = ws.cell(row=row_num+1, column=7)
        cell.value = header4.upper()
        cell.font = Font(name='Times New Roman', bold=True, size=font_size)
        row_num += 1

    headers5 = ["2"]

    for col_num5, header5 in enumerate(headers5, start=1):
        cell = ws.cell(row=row_num, column=1)
        cell.value = header5.upper()
        cell.font = Font(name='Times New Roman',size=font_size)
        row_num += 1

    aggregated_data_expense = {}
    total_sum_expense = 0.0
    for item in data:
        if item.get('grbId') == 2:
            katFgr = item.get('katFgr', '')
            katFgrName = item.get('katFgrName', '')
            klsFpgrName = item.get('klsFpgrName', '')
            sumrg_str = item.get('sumrg', '0').replace(',', '.')
            try:
                sumrg = round(float(sumrg_str) / 1_000_000, 1) if sumrg_str else 0.0
                total_sum_expense += sumrg
            except ValueError:
                sumrg = 0.0

            if katFgr not in aggregated_data_expense:
                aggregated_data_expense[katFgr] = {
                    'katFgrName': katFgrName,
                    'totalSum': 0.0,
                    'klsFpgrNames': {}
                }

            if klsFpgrName in aggregated_data_expense[katFgr]['klsFpgrNames']:
                aggregated_data_expense[katFgr]['klsFpgrNames'][klsFpgrName] += sumrg
            else:
                aggregated_data_expense[katFgr]['klsFpgrNames'][klsFpgrName] = sumrg

            aggregated_data_expense[katFgr]['totalSum'] += sumrg

    sorted_data_expense = dict(sorted(aggregated_data_expense.items()))

    total_sum_cell_expense = ws.cell(row=row_num - 1, column=8)
    total_sum_cell_expense.value = total_sum_expense
    total_sum_cell_expense.font = Font(name='Times New Roman', bold=True, size=font_size)

    out_counter = 1
    in_counter = 1
    for katFgr, katFgr_data in sorted_data_expense.items():
        cell0 = ws.cell(row=row_num, column=3)
        cell1 = ws.cell(row=row_num, column=7)
        cell2 = ws.cell(row=row_num, column=8)
        cell0.value = out_counter
        cell1.value = katFgr_data['katFgrName'].upper()
        cell2.value = katFgr_data['totalSum']
        cell0.font = Font(name='Times New Roman', size=font_size)
        cell1.font = Font(name='Times New Roman', bold=True, size=font_size)
        cell2.font = Font(name='Times New Roman', bold=True, size=font_size)
        row_num += 1
        out_counter += 1

        for klsFpgrName, sumrg in katFgr_data['klsFpgrNames'].items():
            cell0 = ws.cell(row=row_num, column=5)
            cell1 = ws.cell(row=row_num, column=7)
            cell2 = ws.cell(row=row_num, column=8)
            cell0.value = in_counter
            cell1.value = klsFpgrName
            cell2.value = sumrg
            cell0.font = Font(name='Times New Roman', size=font_size2)
            cell1.font = Font(name='Times New Roman', size=font_size2)
            cell2.font = Font(name='Times New Roman', size=font_size)
            cell1.alignment = Alignment(wrap_text=True)
            row_num += 1
            in_counter += 1

    wb.save(output_excel)

@app.route('/state')
def state():
    regions = [
        {"name": "Весь Казахстан", "kato": "00"},
        {"name": "Г. Астана", "kato": f'71'},
        {"name": "Алматинская область", "kato": f'19'},
        {"name": "Г. Шымкент", "kato": '79'},
        {"name": "Акмолинская область", "kato": '11'},
        {"name": "Актюбинская Область", "kato": '15'},
        {"name": "Г. Алматы", "kato": '75'},
        {"name": "Атырауская Область", "kato": '23'},
        {"name": "Восточно-Казахстанская Область", "kato": '63'},
        {"name": "Жамбылская Область", "kato": '31'},
        {"name": "Западно-Казахстанская Область", "kato": '27'},
        {"name": "Карагандинская Область", "kato": '35'},
        {"name": "Кызылординская Область", "kato": '43'},
        {"name": "Костанайская Область", "kato": '39'},
        {"name": "Мангистауская Область", "kato": '47'},
        {"name": "Павлодарская Область", "kato": '55'},
        {"name": "Северо-Казахстанская Область", "kato": '59'},
        {"name": "Туркестанская Область", "kato": '61'},
        {"name": "Абайская область", "kato": '10'},
        {"name": "Область Жетису", "kato": '33'},
        {"name": "Область Улытау", "kato": '62'},
    ]

    dates = [
        {"date": "2022", "file": "json_data/all_data_2022.json"},
        {"date": "2023", "file": "json_data/all_data_2023.json"},
        {"date": "До января", "file": "json_data/all_data_january.json"},
        {"date": "До февраля", "file": "json_data/all_data_february.json"},
        {"date": "До марта", "file": "json_data/all_data_march.json"},
        {"date": "До апреля", "file": "json_data/all_data_april.json"},
        {"date": "До мая", "file": "json_data/all_data_may.json"},
        {"date": "До июня", "file": "json_data/all_data_june.json"},
    ]

    return render_template('state.html', regions=regions, dates=dates)


@app.route('/generate_report', methods=['POST'])
def generate_report():
    kato = request.form.get('region')
    region_name = request.form.get('region_name')
    file_name = request.form.get('file')

    date_1 = request.form.get('date')
    dates = [
        {"date": "2022", "file": "json_data/all_data_2022.json"},
        {"date": "2023", "file": "json_data/all_data_2023.json"},
        {"date": "До января", "file": "json_data/all_data_january.json"},
        {"date": "До февраля", "file": "json_data/all_data_february.json"},
        {"date": "До марта", "file": "json_data/all_data_march.json"},
        {"date": "До апреля", "file": "json_data/all_data_april.json"},
        {"date": "До мая", "file": "json_data/all_data_may.json"},
        {"date": "До июня", "file": "json_data/all_data_june.json"},
    ]

    date_name = ""
    for i in dates:
        if date_1 == i["file"]:
            date_name = i["date"]
            break

    with open(file_name, 'r', encoding='utf-8') as file:
        data = json.load(file)

    filtered_data = [item for item in data if item.get('kato')[:2] == kato and item.get('adtype') != 3 and item.get('adtype') != 4]
    output_excel = f'{region_name}_{date_name}.xlsx'
    create_table_in_excel(filtered_data, region_name, output_excel)

    return send_file(output_excel, as_attachment=True, download_name=output_excel)

def create_table_gos(data, output_excel_gos):


    wb = Workbook()
    ws = wb.active
    ws.title = "Data"


    font_size = 9
    font_size2 = 7

    ws.column_dimensions['A'].width = 12.5
    ws.column_dimensions['B'].width = 12.5
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 5

    for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='center')
            cell.font = Font(name='Times New Roman', size=font_size)

    ws.merge_cells('A1:C1')
    ws['A1'].value = ''
    ws['A1'].font = Font(name='Times New Roman', bold=True)
    ws.merge_cells('A2:C2')
    ws['A2'].value = 'ИCПОЛНЕНИЕ ГОСУДАРСТВЕННОГО БЮДЖЕТА ПО'
    ws['A2'].font = Font(name='Times New Roman', bold=True)
    ws.merge_cells('A3:C3')
    ws['A3'].value = 'ЭКОНОМИЧЕСКОЙ КЛАССИФИКАЦИИ РАСХОДОВ'
    ws['A3'].font = Font(name='Times New Roman', bold=True)

    ws.merge_cells('A4:C4')
    ws['D4'].value = '(млн.тенге)'
    ws['D4'].font = Font(name='Times New Roman', size=font_size)

    ws.merge_cells('A5:A6')
    ws['A5'].value = 'Категория'
    ws['A5'].font = Font(name='Times New Roman', size=font_size)
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 80

    ws.merge_cells('B5:B6')
    ws['B5'].value = 'Подкласс'
    ws['B5'].font = Font(name='Times New Roman', size=font_size)

    ws.merge_cells('C5:C6')
    ws['C5'].value = 'Наименование'
    ws['C5'].font = Font(name='Times New Roman', size=font_size)
    ws['C5'].alignment = Alignment(vertical='center', horizontal='center')

    ws.merge_cells('E5:E6')

    headers3 = ["I. Доходы"]

    for col_num3, header3 in enumerate(headers3, start=1):
        cell = ws.cell(row=8, column=3)
        cell.value = header3.upper()
        cell.font = Font(name='Times New Roman', bold=True, size=font_size)

    aggregated_data = {}
    total_sum = 0.0
    for item in data:
        if item.get('grbId') == 1:
            katFgr = item.get('katFgr', '')
            katFgrName = item.get('katFgrName', '')
            klsFpgrName = item.get('klsFpgrName', '')
            sumrg_str = item.get('sumrg', '0').replace(',', '.')
            try:
                sumrg = float(sumrg_str)/1_000_000 if sumrg_str else 0.0
            except ValueError:
                sumrg = 0.0

            if katFgr == "4":
                if item.get('klsFpgr') != "4":
                    continue

            if katFgr not in aggregated_data:
                aggregated_data[katFgr] = {
                    'katFgrName': katFgrName,
                    'totalSum': 0.0,
                    'klsFpgrNames': {}
                }

            if klsFpgrName in aggregated_data[katFgr]['klsFpgrNames']:
                aggregated_data[katFgr]['klsFpgrNames'][klsFpgrName] += sumrg
            else:
                aggregated_data[katFgr]['klsFpgrNames'][klsFpgrName] = sumrg

            aggregated_data[katFgr]['totalSum'] += sumrg
            total_sum += sumrg


    sorted_data = dict(sorted(aggregated_data.items()))

    row_num = 9
    counter = 1
    counter_kls = 1
    for katFgr, katFgr_data in sorted_data.items():
        cell0 = ws.cell(row_num, column=1)
        cell1 = ws.cell(row=row_num, column=3)
        cell2 = ws.cell(row=row_num, column=4)
        cell0.value = counter
        cell1.value = katFgr_data['katFgrName'].upper()
        cell2.value = katFgr_data['totalSum']
        cell1.font = Font(name='Times New Roman', bold=True, size=font_size)
        cell2.font = Font(name='Times New Roman', bold=True, size=font_size)
        counter += 1
        row_num += 1

        for klsFpgrName, sumrg in katFgr_data['klsFpgrNames'].items():
            cell0 = ws.cell(row=row_num, column=2)
            cell1 = ws.cell(row=row_num, column=3)
            cell2 = ws.cell(row=row_num, column=4)
            cell0.value = counter_kls
            cell1.value = klsFpgrName
            cell2.value = sumrg
            cell1.font = Font(name='Times New Roman', size=font_size2)
            cell2.font = Font(name='Times New Roman', size=font_size)
            cell1.alignment = Alignment(wrap_text=True)
            row_num += 1
            counter_kls += 1

    total_sum_cell = ws.cell(row=8, column=4)
    total_sum_cell.value = total_sum
    total_sum_cell.font = Font(name='Times New Roman', bold=True, size=font_size)

    row_num += 0

    headers4 = ["II. Затраты"]

    for col_num4, header4 in enumerate(headers4, start=1):
        cell = ws.cell(row=row_num, column=3)
        cell.value = header4.upper()
        cell.font = Font(name='Times New    Roman', bold=True, size=font_size)
        row_num += 1

    aggregated_data_expense = {}
    total_sum_expense = 0.0
    for item in data:
        if item.get('grbId') == 2:
            katFgr = item.get('katFgr', '')
            katFgrName = item.get('katFgrName', '')
            klsFpgrName = item.get('klsFpgrName', '')
            sumrg_str = item.get('sumrg', '0').replace(',', '.')
            try:
                sumrg = round(float(sumrg_str) / 1_000_000, 1) if sumrg_str else 0.0
                total_sum_expense += sumrg
            except ValueError:
                sumrg = 0.0

            if katFgr not in aggregated_data_expense:
                aggregated_data_expense[katFgr] = {
                    'katFgrName': katFgrName,
                    'totalSum': 0.0,
                    'klsFpgrNames': {}
                }

            if klsFpgrName in aggregated_data_expense[katFgr]['klsFpgrNames']:
                aggregated_data_expense[katFgr]['klsFpgrNames'][klsFpgrName] += sumrg
            else:
                aggregated_data_expense[katFgr]['klsFpgrNames'][klsFpgrName] = sumrg

            aggregated_data_expense[katFgr]['totalSum'] += sumrg

    sorted_data_expense = dict(sorted(aggregated_data_expense.items()))

    total_sum_cell_expense = ws.cell(row=row_num - 1, column=4)
    total_sum_cell_expense.value = total_sum_expense
    total_sum_cell_expense.font = Font(name='Times New Roman', bold=True, size=font_size)

    out_counter = 1
    in_counter = 1
    for katFgr, katFgr_data in sorted_data_expense.items():
        cell0 = ws.cell(row=row_num, column=1)
        cell1 = ws.cell(row=row_num, column=3)
        cell2 = ws.cell(row=row_num, column=4)
        cell0.value = out_counter
        cell1.value = katFgr_data['katFgrName'].upper()
        cell2.value = katFgr_data['totalSum']
        cell0.font = Font(name='Times New Roman', size=font_size)
        cell1.font = Font(name='Times New Roman', bold=True, size=font_size)
        cell2.font = Font(name='Times New Roman', bold=True, size=font_size)
        row_num += 1
        out_counter += 1

        for klsFpgrName, sumrg in katFgr_data['klsFpgrNames'].items():
            cell0 = ws.cell(row=row_num, column=2)
            cell1 = ws.cell(row=row_num, column=3)
            cell2 = ws.cell(row=row_num, column=4)
            cell0.value = in_counter
            cell1.value = klsFpgrName
            cell2.value = sumrg
            cell0.font = Font(name='Times New Roman', size=font_size2)
            cell1.font = Font(name='Times New Roman', size=font_size2)
            cell2.font = Font(name='Times New Roman', size=font_size)
            cell1.alignment = Alignment(wrap_text=True)
            row_num += 1
            in_counter += 1

    row_num += 0

    headers4 = ["III. ЧИСТОЕ БЮДЖЕТНОЕ КРЕДИТОВАНИЕ"]

    for col_num4, header4 in enumerate(headers4, start=1):
        cell = ws.cell(row=row_num, column=3)
        cell.value = header4.upper()
        cell.font = Font(name='Times New    Roman', bold=True, size=font_size)
        row_num += 1

    aggregated_data_expense = {}
    total_sum_expense = 0.0
    for item in data:
        if item.get('grbId') == 3:
            katFgr = item.get('katFgr', '')
            katFgrName = item.get('katFgrName', '')
            klsFpgrName = item.get('klsFpgrName', '')
            sumrg_str = item.get('sumrg', '0').replace(',', '.')
            try:
                sumrg = round(float(sumrg_str) / 1_000_000, 1) if sumrg_str else 0.0
                total_sum_expense += sumrg
            except ValueError:
                sumrg = 0.0

            if katFgr not in aggregated_data_expense:
                aggregated_data_expense[katFgr] = {
                    'katFgrName': katFgrName,
                    'totalSum': 0.0,
                    'klsFpgrNames': {}
                }

            if klsFpgrName in aggregated_data_expense[katFgr]['klsFpgrNames']:
                aggregated_data_expense[katFgr]['klsFpgrNames'][klsFpgrName] += sumrg
            else:
                aggregated_data_expense[katFgr]['klsFpgrNames'][klsFpgrName] = sumrg

            aggregated_data_expense[katFgr]['totalSum'] += sumrg

    sorted_data_expense = dict(sorted(aggregated_data_expense.items()))

    total_sum_cell_expense = ws.cell(row=row_num - 1, column=4)
    total_sum_cell_expense.value = total_sum_expense
    total_sum_cell_expense.font = Font(name='Times New Roman', bold=True, size=font_size)

    out_counter = 1
    in_counter = 1
    for katFgr, katFgr_data in sorted_data_expense.items():
        cell0 = ws.cell(row=row_num, column=1)
        cell1 = ws.cell(row=row_num, column=3)
        cell2 = ws.cell(row=row_num, column=4)
        cell0.value = out_counter
        cell1.value = katFgr_data['katFgrName'].upper()
        cell2.value = katFgr_data['totalSum']
        cell0.font = Font(name='Times New Roman', size=font_size)
        cell1.font = Font(name='Times New Roman', bold=True, size=font_size)
        cell2.font = Font(name='Times New Roman', bold=True, size=font_size)
        row_num += 1
        out_counter += 1

        for klsFpgrName, sumrg in katFgr_data['klsFpgrNames'].items():
            cell0 = ws.cell(row=row_num, column=2)
            cell1 = ws.cell(row=row_num, column=3)
            cell2 = ws.cell(row=row_num, column=4)
            cell0.value = in_counter
            cell1.value = klsFpgrName
            cell2.value = sumrg
            cell0.font = Font(name='Times New Roman', size=font_size2)
            cell1.font = Font(name='Times New Roman', size=font_size2)
            cell2.font = Font(name='Times New Roman', size=font_size)
            cell1.alignment = Alignment(wrap_text=True)
            row_num += 1
            in_counter += 1

    wb.save(output_excel_gos)

@app.route('/generate_gos', methods=['POST'])
def generate_gos():
    file_name = request.form.get('file_gos')

    date_1 = request.form.get('date1')
    dates = [
        {"date": "2022", "file": "json_data/all_data_2022.json"},
        {"date": "2023", "file": "json_data/all_data_2023.json"},
        {"date": "На январь", "file": "json_data/all_data_january.json"},
        {"date": "На февраль", "file": "json_data/all_data_february.json"},
        {"date": "На март", "file": "json_data/all_data_march.json"},
        {"date": "На апрель", "file": "json_data/all_data_april.json"},
        {"date": "На май", "file": "json_data/all_data_may.json"},
        {"date": "На июнь", "file": "json_data/all_data_june.json"},
    ]

    date_name = ""
    for i in dates:
        if date_1 == i["file"]:
            date_name = i["date"]
            break

    with open(file_name, 'r', encoding='utf-8') as file:
        data = json.load(file)

    filtered_data = [item for item in data if item.get('adtype') != 3 and item.get('adtype') != 4]
    output_excel_gos = f'ГосБюджет_{date_name}.xlsx'
    create_table_gos(filtered_data, output_excel_gos)

    return send_file(output_excel_gos, as_attachment=True, download_name=output_excel_gos)

if __name__ == '__main__':
    app.run(debug=True)
